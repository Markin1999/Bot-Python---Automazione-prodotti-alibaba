from openpyxl import Workbook, load_workbook
import os

file_path = "file/macro.xlsx"

def modificaExcel(cella, testo):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        wb.save(file_path)
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws[cella] = testo
    wb.save(file_path)
    print(f"modificaExcel.py/ ✅ La cella {cella} è stata modificata correttamente in macro.xlsx")

    