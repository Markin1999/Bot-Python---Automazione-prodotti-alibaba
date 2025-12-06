from openpyxl import Workbook, load_workbook
import os

file_path = "file/macro.xlsx"

def scritturaExcel(wb, ws):

    ws["A1"] = "NumeroProcesso"
    ws["B1"] = "Processi"
    ws["C1"] = "Ricerca"
    ws["D1"] = "ParoleChiave"
    ws["E1"] = "Nome"


    prodotti = [

       "bamboo toilet paper 5 ply 50m bamboo core 100 percent bamboo pulp plastic free FSC Ecolabel OEM",
        "bamboo jumbo tissue roll large and mini jumbo 100 percent bamboo pulp plastic free FSC OEM",
        "bamboo paper hand towels roll or folded 100 percent bamboo pulp plastic free FSC OEM",
        "A4 copy paper 80gsm 100 percent bamboo pulp plastic free FSC Ecolabel OEM custom logo",
        "notebooks and bloc-notes bamboo paper pages kraft cover plastic free FSC custom logo",
        "facial tissues 100 percent bamboo pulp pocket or box plastic free FSC Ecolabel OEM",
        "kraft paper tape water-activated gummed biodegradable plastic free FSC custom print",
        "bamboo kraft recycled paper packaging boxes and mailers plastic free FSC custom branding"
   ]

    parole_chiave = [
        "packaging sostenibile",
        "imballaggio sostenibile",
        "packaging ecologico",
        "imballaggio ecologico",
        "packaging biodegradabile",
        "imballaggio biodegradabile",
        "packaging compostabile",
        "imballaggio compostabile",
        "packaging riciclabile",
        "imballaggio riciclabile",
        "carta kraft",
        "carta riciclata",
        "cellulosa di bambù",
        "fibra di bambù",
        "materiale riciclato",
        "materiale ecologico",
        "materiale sostenibile",
        "bambù naturale",
        "cartone riciclato",
        "eco friendly",
        "prodotto ecologico",
        "scatola ecologica",
        "scatola sostenibile",
        "packaging personalizzato",
        "imballaggio personalizzato",
        "stampa personalizzata",
        "etichetta ecologica",
        "busta compostabile",
        "sacchetto biodegradabile",
        "spedizione campioni"
    ]

    nomi = [
        "carta",
        "carta",
        "carta",
        "carta",
        "carta",
        "carta",
        "carta",
        "carta"
    ]



    righe_max = max(len(prodotti), len(parole_chiave), len(nomi))

    for i in range(2, righe_max + 2):
        prodotto = prodotti[i - 2] if i - 2 < len(prodotti) else ""
        parola = parole_chiave[i - 2] if i - 2 < len(parole_chiave) else ""
        nome = nomi[i - 2] if i - 2 < len(nomi) else ""
        ws[f"B{i}"] = str(i - 1)
        ws[f"C{i}"] = prodotto
        ws[f"D{i}"] = parola
        ws[f"E{i}"] = nome



    wb.save(file_path)
    print("✅ Tutti i prodotti sono stati scritti correttamente in macro.xlsx")


if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    wb.save(file_path)
else:
    wb = load_workbook(file_path)
    ws = wb.active

scritturaExcel(wb, ws)
